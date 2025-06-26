from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
import os
from typing import Optional
from datetime import datetime, timedelta, date
from models.daily_batch import DailyBatch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
from database import get_db
import pandas as pd
from io import BytesIO
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import and_
import models 

def get_daily_batches_by_date_range(db: Session, start_date: date, end_date: date):
    return db.query(models.DailyBatch).filter(
        and_(models.DailyBatch.batch_date >= start_date,
             models.DailyBatch.batch_date <= end_date)
    ).all()

router = APIRouter(
    prefix="/reports",
    tags=["reports"],
)

@router.get("/daily-report")
def get_daily_report(start_date: Optional[str] = None, end_date: Optional[str] = None, db: Session = Depends(get_db)):
    if not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Both start_date and end_date parameters are required")

    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Please use YYYY-MM-DD for both start_date and end_date")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date cannot be after the end date")

    daily_data = get_daily_batches_by_date_range(db, start_date=start_date_obj, end_date=end_date_obj)

    if not daily_data:
        raise HTTPException(status_code=200, detail=f"No data found between {start_date} and {end_date}")
    
    report_list = [item.__dict__ for item in daily_data]
    for item in report_list:
        del item['_sa_instance_state']  # Remove SQLAlchemy internal attribute

    # Create a pandas DataFrame
    df = pd.DataFrame(report_list)

    # Create an in-memory Excel file
    excel_file = BytesIO()
    df.to_excel(excel_file, index=False, sheet_name='Daily Report')
    excel_file.seek(0)

    # Prepare the response
    headers = {
        'Content-Disposition': 'attachment; filename="daily_report.xlsx"'
    }

    return StreamingResponse(excel_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.get("/snapshot")
def get_snapshot(start_date: str, end_date: str, batch_id: Optional[int] = None, db: Session = Depends(get_db)):
    """
    Get a snapshot of batches between the specified start_date and end_date with total_eggs count.
    Optionally filter by batch_id. Returns all rows for the batch_id if provided.
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Please use YYYY-MM-DD for both start_date and end_date")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date cannot be after the end date")

    # Query DailyBatch for the specified date range, ordered by batch_id and batch_date
    query = db.query(DailyBatch).filter(
        DailyBatch.batch_date >= start_date_obj,
        DailyBatch.batch_date <= end_date_obj
    )

    if batch_id is not None:
        query = query.filter(DailyBatch.batch_id == batch_id)

    query = query.order_by(DailyBatch.batch_id.asc(), DailyBatch.batch_date.asc())
    daily_batches = query.all()

    # Serialize the data, including total_eggs
    result = [
        {
            "batch_id": batch.batch_id,
            "batch_no": batch.batch_no,
            "batch_date": batch.batch_date.strftime("%d-%m-%Y"),
            "shed_no": batch.shed_no,
            "age": batch.age,
            "opening_count": batch.opening_count,
            "mortality": batch.mortality,
            "culls": batch.culls,
            "closing_count": batch.closing_count,
            "table_eggs": batch.table_eggs,
            "jumbo": batch.jumbo,
            "cr": batch.cr,
            "total_eggs": batch.total_eggs,  # Computed property
            "hd": batch.hd,  # Computed property
            "standard_hen_day_percentage": float(batch.standard_hen_day_percentage) if batch.standard_hen_day_percentage is not None else batch.standard_hen_day_percentage,
        }
        for batch in daily_batches
    ]

    return JSONResponse(content=result)

def write_daily_report_excel(batches, report_date=None, file_path=None):
    if report_date is None:
        report_date = date.today()
    if file_path is None:
        file_path = "daily_report_combined.xlsx"  # Using a single combined file

    report_date_str = report_date.strftime("%d-%m-%Y")
    header1 = ["DATE", report_date_str, "", "", "DAILY REPORT", "", "", "", "", "", "", "","", ""]
    # Adjust header1 to match the number of columns in header2
    header2 = ["BATCH", "SHED", "AGE", "OPEN", "MORT", "CULLS", "CLOSING", "TABLE", "JUMBO", "CR", "TOTAL", "HD%", "STD"]

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

    # Find if today's report already exists and remove it
    found_row = None
    for i in range(ws.max_row, 0, -1):  # Start from last row to first
        row = [cell.value for cell in ws[i]]
        if row[0] == "DATE" and row[1] == report_date_str:
            found_row = i
            break


    if found_row:
        # Remove existing block (today's DATE to TOTAL row)
        end_row = found_row
        while end_row <= ws.max_row:
            if ws.cell(row=end_row, column=1).value == "TOTAL":
                break
            end_row += 1
        ws.delete_rows(found_row, end_row - found_row + 1)

    # Append data to bottom
    # Check if the last row is empty and add a blank row for spacing
    last_row_values = [cell.value for cell in ws[ws.max_row]]
    if any(last_row_values):  # if the last row is not completely empty
        ws.append([])         # add a blank row for spacing
        start_row = ws.max_row + 2
    else:
        start_row = ws.max_row + 1

    # Now append header and data
    ws.append(header1)
    ws.append(header2)
    
    # Write batch rows
    total_open = total_mort = total_culls = total_closing = total_table = total_jumbo = total_cr = total_total = total_hd_percent_dummy = total_std_dummy = 0
    for batch in batches:
        row = [
            batch.batch_no,
            batch.shed_no,
            batch.age,
            batch.opening_count,
            batch.mortality,
            batch.culls,
            batch.closing_count,
            getattr(batch, "table_eggs", 0),
            getattr(batch, "jumbo", 0),
            getattr(batch, "cr", 0),
            getattr(batch, "total", 0),
            round(batch.closing_count / max(1, batch.opening_count), 4), # Dummy HD% calculation
            round(float(batch.age) * 2.5, 1) if batch.age else round(float(batch.age) * 2.5, 1) # Dummy STD calculation
        ]
        ws.append(row)
        total_open += batch.opening_count
        total_mort += batch.mortality
        total_culls += batch.culls
        total_closing += batch.closing_count
        total_table += getattr(batch, "table_eggs", 0)
        total_jumbo += getattr(batch, "jumbo", 0)
        total_cr += getattr(batch, "cr", 0)
        total_total += getattr(batch, "total", 0)
        # Accumulate dummy values for totals
        total_hd_percent_dummy += round(batch.closing_count / max(1, batch.opening_count), 4)
        total_std_dummy += round(float(batch.age) * 2.5, 1)
        

    # Write totals row
    ws.append([
        "TOTAL", "", 0, total_open, total_mort, total_culls, total_closing,
        total_table, total_jumbo, total_cr, total_total,
         round(total_hd_percent_dummy / len(batches) if batches else 0, 4), # Dummy average HD%
         round(total_std_dummy / len(batches) if batches else 0, 1) # Dummy average STD
    ])

    # Apply formatting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_red_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid") # Closer to the image's orange-red
    bold_font_black = Font(bold=True, color="000000") # Black font for yellow fill
    bold_font_white = Font(bold=True, color="FFFFFF") # White font for red/orange-red fill
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    bold_font = Font(bold=True)

     # Formatting for header1 (DATE and DAILY REPORT)
    date_cell = ws.cell(row=start_row, column=1)
    date_cell.fill = yellow_fill
    date_cell.font = bold_font_black

    report_date_cell = ws.cell(row=start_row, column=2)
    report_date_cell.fill = yellow_fill
    report_date_cell.font = bold_font_black

    daily_report_cell = ws.cell(row=start_row, column=5) # Column E (5th column) for "DAILY REPORT"
    daily_report_cell.value = "DAILY REPORT"
    daily_report_cell.fill = red_fill
    daily_report_cell.font = bold_font_white
    # Merge cells for "DAILY REPORT"
    ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=len(header2))
    daily_report_cell.alignment = Alignment(horizontal='center', vertical='center')


    # Formatting for header2 (BATCH, SHED, etc.)
    for col_idx, cell in enumerate(ws[start_row + 1]):
        cell.fill = orange_red_fill
        cell.font = bold_font_white
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10 # Adjust width as needed

    # Formatting for Total row
    for cell in ws[ws.max_row]:
        cell.fill = yellow_fill
        cell.font = bold_font_black

    wb.save(file_path)
    return file_path


