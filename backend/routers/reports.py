from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
import os
from typing import Optional, Dict
from datetime import datetime, timedelta, date
from models.daily_batch import DailyBatch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
from database import get_db
import pandas as pd
from io import BytesIO
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import and_
import models
from models.batch import Batch
from utils.tenancy import get_tenant_id
from models.bovanswhitelayerperformance import BovansWhiteLayerPerformance

def get_daily_batches_by_date_range(db: Session, start_date: date, end_date: date, tenant_id: str):
    return db.query(models.DailyBatch).filter(
        and_(models.DailyBatch.batch_date >= start_date,
             models.DailyBatch.batch_date <= end_date,
             models.DailyBatch.tenant_id == tenant_id)
    ).all()

router = APIRouter(
    prefix="/reports",
    tags=["reports"],
)

def _calculate_cumulative_report(db: Session, batch_id: int, current_week: int, hen_housing: int, current_summary: dict, tenant_id: str):
    """Calculate cumulative report data"""
    
    # Calculate cumulative feed from week 18 to current week
    cum_feed_total = 0
    for week_num in range(18, current_week + 1):
        week_batches = db.query(DailyBatch).filter(
            DailyBatch.batch_id == batch_id,
            DailyBatch.age >= f"{week_num}.1",
            DailyBatch.age <= f"{week_num}.7",
            DailyBatch.tenant_id == tenant_id
        ).all()
        
        week_feed = sum(float(batch.feed_in_kg) for batch in week_batches if batch.feed_in_kg is not None)
        cum_feed_total += week_feed
    
    # Get Bovans standard data for current week
    bovans_data = db.query(BovansWhiteLayerPerformance).filter(
        BovansWhiteLayerPerformance.age_weeks == current_week,
        BovansWhiteLayerPerformance.tenant_id == tenant_id
    ).first()
    
    # Section 1: Feed data
    section1 = {
        "cum_feed": {
            "cum": float(cum_feed_total),
            "actual": round(cum_feed_total / hen_housing, 4) if hen_housing > 0 else 0,
            "standard": float(bovans_data.feed_intake_cum_kg) if bovans_data and bovans_data.feed_intake_cum_kg else 0,
            "diff": 0  # Will calculate after
        },
        "weekly_feed": {
            "cum": float(current_summary["actual_feed_consumed"]),
            "actual": round(current_summary["actual_feed_consumed"] / hen_housing, 4) if hen_housing > 0 else 0,
            "standard": round((float(bovans_data.feed_intake_per_day_g) * 7 / 1000), 4) if bovans_data and bovans_data.feed_intake_per_day_g else 0,
            "diff": 0  # Will calculate after
        }
    }
    
    # Calculate diff for section 1
    section1["cum_feed"]["diff"] = round(section1["cum_feed"]["actual"] - section1["cum_feed"]["standard"], 4)
    section1["weekly_feed"]["diff"] = round(section1["weekly_feed"]["actual"] - section1["weekly_feed"]["standard"], 4)
    
    # Section 2: Performance data
    section2 = {
        "livability": {
            "actual": round((current_summary["closing_count"] / hen_housing) * 100, 2) if hen_housing > 0 else 0,
            "standard": float(bovans_data.livability_percent) if bovans_data and bovans_data.livability_percent else 0,
            "diff": 0  # Will calculate after
        },
        "feed_grams": {
            "actual": round((current_summary["actual_feed_consumed"] * 1000) / hen_housing, 2) if hen_housing > 0 else 0,
            "standard": float(bovans_data.feed_intake_per_day_g) if bovans_data and bovans_data.feed_intake_per_day_g else 0,
            "diff": 0  # Will calculate after
        }
    }
    
    # Calculate diff for section 2
    section2["livability"]["diff"] = round(section2["livability"]["actual"] - section2["livability"]["standard"], 2)
    section2["feed_grams"]["diff"] = round(section2["feed_grams"]["actual"] - section2["feed_grams"]["standard"], 2)
    
    return {
        "section1": section1,
        "section2": section2
    }

def _calculate_summary(batches: list[DailyBatch], query_start_date: date, query_end_date: date, is_single_batch: bool):
    """Helper function to calculate summary statistics for a list of daily batches."""
    if not batches:
        return None

    summary_opening_count = 0
    summary_closing_count = 0

    from itertools import groupby

    # The caller must sort batches by batch_id, then date.
    grouped_batches = groupby(batches, key=lambda b: b.batch_id)

    if is_single_batch:
        # For a single batch, take the opening count of the first day and closing count of the last day.
        # The 'batches' list is already sorted by date.
        summary_opening_count = batches[0].opening_count
        summary_closing_count = batches[-1].closing_count
    else:
        # For multiple batches (consolidation), sum the opening count of the first day and closing count of the last day for each batch.

        summary_opening_count = 0
        summary_closing_count = 0

        for _, group in grouped_batches:
            batch_records = list(group)
            if batch_records:
                summary_opening_count += batch_records[0].opening_count
                summary_closing_count += batch_records[-1].closing_count

    # Calculate highest age across all batches and all dates in the provided list
    highest_age = 0.0
    for b in batches:
        try:
            age_val = float(b.age)
            if age_val > highest_age:
                highest_age = age_val
        except (ValueError, TypeError):
            continue # Ignore invalid age values

    # Sum totals
    total_mortality = sum(b.mortality for b in batches)
    total_culls = sum(b.culls for b in batches)
    total_table_eggs = sum(b.table_eggs for b in batches)
    total_jumbo = sum(b.jumbo for b in batches if b.jumbo is not None)
    total_cr = sum(b.cr for b in batches)
    total_eggs = total_table_eggs + total_jumbo + total_cr

    # Filter for layer batches (age >= 18 weeks) for HD calculations
    layer_batches = []
    for b in batches:
        try:
            # Only include batches that are 18 weeks or older
            if float(b.age) >= 18:
                layer_batches.append(b)
        except (ValueError, TypeError):
            # Ignore if age is not a valid number
            continue

    # Calculate averages only on layer batches
    if layer_batches:
        # Ensure hd and standard_hen_day_percentage are not None before summing
        hd_values = [b.hd for b in layer_batches if b.hd is not None]
        avg_hd = sum(hd_values) / len(hd_values) if hd_values else 0

        standard_hd_values = [float(b.standard_hen_day_percentage) for b in layer_batches if b.standard_hen_day_percentage is not None]
        avg_standard_hd = sum(standard_hd_values) / len(standard_hd_values) if standard_hd_values else 0
    else:
        avg_hd = 0
        avg_standard_hd = 0

    return {
        "opening_count": summary_opening_count,
        "mortality": total_mortality,
        "culls": total_culls,
        "closing_count": summary_closing_count,
        "table_eggs": total_table_eggs,
        "jumbo": total_jumbo,
        "cr": total_cr,
        "total_eggs": total_eggs,
        "hd": round(avg_hd, 4),
        "standard_hen_day_percentage": round(avg_standard_hd, 4),
        "highest_age": round(highest_age, 1),
    }


@router.get("/weekly-layer-report")
def get_weekly_layer_report(
    batch_id: int,
    week: int,  # e.g., 18 for week 18.1-18.7
    db: Session = Depends(get_db),
    tenant_id: str = Depends(get_tenant_id)
):
    """Get weekly report for layer birds based on age range (e.g., week 18 = 18.1 to 18.7)"""
    
    # Calculate age range for the week
    start_age = f"{week}.1"
    end_age = f"{week}.7"
    
    # Get daily batches for the specified age range
    daily_batches = db.query(DailyBatch).filter(
        DailyBatch.batch_id == batch_id,
        DailyBatch.age >= start_age,
        DailyBatch.age <= end_age,
        DailyBatch.tenant_id == tenant_id
    ).order_by(DailyBatch.batch_date.asc()).all()
    
    if not daily_batches:
        raise HTTPException(status_code=404, detail=f"No data found for batch {batch_id} at week {week}")
    
    # Get hen housing (closing count at age 17.7)
    hen_housing_record = db.query(DailyBatch).filter(
        DailyBatch.batch_id == batch_id,
        DailyBatch.age == "17.7",
        DailyBatch.tenant_id == tenant_id
    ).first()
    
    hen_housing = hen_housing_record.closing_count if hen_housing_record else 0
    
    # Calculate summary using existing logic
    start_date = daily_batches[0].batch_date
    end_date = daily_batches[-1].batch_date
    summary_data = _calculate_summary(daily_batches, start_date, end_date, is_single_batch=True)
    
    if summary_data:
        total_actual_feed_consumed = sum(record.feed_in_kg for record in daily_batches if record.feed_in_kg is not None)
        total_standard_feed_consumption = sum(record.standard_feed_in_kg * record.opening_count for record in daily_batches if record.standard_feed_in_kg is not None and record.opening_count is not None)
        
        summary_data["actual_feed_consumed"] = total_actual_feed_consumed
        summary_data["standard_feed_consumption"] = total_standard_feed_consumption
        summary_data["hen_housing"] = hen_housing
        
        # Calculate hen housing percentages
        if hen_housing > 0:
            summary_data["opening_percent"] = round((summary_data["opening_count"] * 100) / hen_housing, 2)
            summary_data["mort_percent"] = round((summary_data["mortality"] * 100) / hen_housing, 2)
            summary_data["culls_percent"] = round((summary_data["culls"] * 100) / hen_housing, 2)
            summary_data["closing_percent"] = round((summary_data["closing_count"] * 100) / hen_housing, 2)
            summary_data["feed_per_bird_per_day_grams"] = round((total_actual_feed_consumed * 1000) / (hen_housing * 7), 2)
        else:
            summary_data["opening_percent"] = 0
            summary_data["mort_percent"] = 0
            summary_data["culls_percent"] = 0
            summary_data["closing_percent"] = 0
            summary_data["feed_per_bird_per_day_grams"] = 0
    
    # Prepare detailed results
    detailed_result = []
    for batch in daily_batches:
        actual_feed_consumed = batch.feed_in_kg
        standard_feed_consumption = None
        if batch.standard_feed_in_kg is not None and batch.opening_count is not None:
            standard_feed_consumption = batch.standard_feed_in_kg * batch.opening_count

        detailed_result.append({
            "batch_id": batch.batch_id,
            "batch_no": batch.batch_no,
            "batch_date": batch.batch_date.strftime("%d-%m-%Y"),
            "shed_no": batch.shed_no,
            "age": batch.age,
            "opening_count": batch.opening_count,
            "mortality": batch.mortality,
            "culls": batch.culls,
            "closing_count": batch.closing_count,
            "table_eggs": batch.table_eggs,
            "jumbo": batch.jumbo,
            "cr": batch.cr,
            "total_eggs": batch.total_eggs,
            "batch_type": batch.batch_type,
            "hd": batch.hd,
            "standard_hen_day_percentage": float(batch.standard_hen_day_percentage) if batch.standard_hen_day_percentage is not None else None,
            "actual_feed_consumed": actual_feed_consumed,
            "standard_feed_consumption": standard_feed_consumption,
        })
    
    # Calculate cumulative report
    cumulative_report = _calculate_cumulative_report(db, batch_id, week, hen_housing, summary_data, tenant_id)
    
    return JSONResponse(content={
        "details": detailed_result,
        "summary": summary_data,
        "week": week,
        "age_range": f"{start_age} - {end_age}",
        "hen_housing": hen_housing,
        "cumulative_report": cumulative_report
    })

@router.get("/snapshot")
def get_snapshot(start_date: str, end_date: str, batch_id: Optional[int] = None, db: Session = Depends(get_db), tenant_id: str = Depends(get_tenant_id)):
    """
    Get a snapshot of batches between the specified start_date and end_date.
    If batch_id is provided, returns all rows for that batch.
    If batch_id is not provided, consolidates data into one row per batch.
    """
    try:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Please use YYYY-MM-DD for both start_date and end_date")

    if start_date_obj > end_date_obj:
        raise HTTPException(status_code=400, detail="Start date cannot be after the end date")

    query = db.query(DailyBatch).join(Batch).filter(
        DailyBatch.batch_date >= start_date_obj,
        DailyBatch.batch_date <= end_date_obj,
        Batch.is_active,
        Batch.tenant_id == tenant_id
    )

    summary_data = None
    detailed_result = []

    if batch_id is not None:
        query = query.filter(DailyBatch.batch_id == batch_id)
        daily_batches = query.order_by(DailyBatch.batch_date.asc()).all()
        summary_data = _calculate_summary(daily_batches, start_date_obj, end_date_obj, is_single_batch=True)

        if summary_data:
            total_actual_feed_consumed = sum(record.feed_in_kg for record in daily_batches if record.feed_in_kg is not None)
            total_standard_feed_consumption = sum(record.standard_feed_in_kg * record.opening_count for record in daily_batches if record.standard_feed_in_kg is not None and record.opening_count is not None)
            
            summary_data["actual_feed_consumed"] = total_actual_feed_consumed
            summary_data["standard_feed_consumption"] = total_standard_feed_consumption

        # For a single batch, the details are the daily records
        for batch in daily_batches:
            actual_feed_consumed = batch.feed_in_kg
            standard_feed_consumption = None
            if batch.standard_feed_in_kg is not None and batch.opening_count is not None:
                standard_feed_consumption = batch.standard_feed_in_kg * batch.opening_count

            detailed_result.append({
                "batch_id": batch.batch_id,
                "batch_no": batch.batch_no,
                "batch_date": batch.batch_date.strftime("%d-%m-%Y"),
                "shed_no": batch.shed_no,
                "age": batch.age,
                "opening_count": batch.opening_count,
                "mortality": batch.mortality,
                "culls": batch.culls,
                "closing_count": batch.closing_count,
                "table_eggs": batch.table_eggs,
                "jumbo": batch.jumbo,
                "cr": batch.cr,
                "total_eggs": batch.total_eggs,
                "batch_type": batch.batch_type,
                "hd": batch.hd,
                "standard_hen_day_percentage": float(batch.standard_hen_day_percentage) if batch.standard_hen_day_percentage is not None else None,
                "actual_feed_consumed": actual_feed_consumed,
                "standard_feed_consumption": standard_feed_consumption,
            })

    else:
        # No batch_id provided, so we consolidate.
        all_daily_batches = query.order_by(DailyBatch.batch_id, DailyBatch.batch_date.asc()).all()
        
        # First, calculate the grand summary over all batches
        summary_data = _calculate_summary(all_daily_batches, start_date_obj, end_date_obj, is_single_batch=False)

        if summary_data:
            total_actual_feed_consumed = sum(record.feed_in_kg for record in all_daily_batches if record.feed_in_kg is not None)
            total_standard_feed_consumption = sum(record.standard_feed_in_kg * record.opening_count for record in all_daily_batches if record.standard_feed_in_kg is not None and record.opening_count is not None)
            
            summary_data["actual_feed_consumed"] = total_actual_feed_consumed
            summary_data["standard_feed_consumption"] = total_standard_feed_consumption

        # Group batches by batch_id for detailed consolidation
        from itertools import groupby
        
        grouped_batches = groupby(all_daily_batches, key=lambda b: b.batch_id)

        for b_id, group in grouped_batches:
            batch_records = list(group)
            # Calculate a summary for this specific batch's records
            batch_summary = _calculate_summary(batch_records, batch_records[0].batch_date, batch_records[-1].batch_date, is_single_batch=True)
            
            if batch_summary:
                # Calculate feed consumption for the batch
                total_actual_feed_consumed = sum(record.feed_in_kg for record in batch_records if record.feed_in_kg is not None)
                total_standard_feed_consumption = sum(record.standard_feed_in_kg * record.opening_count for record in batch_records if record.standard_feed_in_kg is not None and record.opening_count is not None)
                
                batch_summary["actual_feed_consumed"] = total_actual_feed_consumed
                batch_summary["standard_feed_consumption"] = total_standard_feed_consumption
                # Add batch-specific info to the summary
                batch_summary["batch_id"] = b_id
                batch_summary["batch_no"] = batch_records[0].batch_no
                batch_summary["shed_no"] = batch_records[0].shed_no
                # Use the batch_type from the last record in the date range for this batch
                batch_summary["batch_type"] = batch_records[-1].batch_type
                detailed_result.append(batch_summary)

        detailed_result.sort(key=lambda x: x.get('batch_no', ''))

    response_content = {
        "details": detailed_result,
        "summary": summary_data
    }

    return JSONResponse(content=response_content)

def write_daily_report_excel(batches, report_date=None, file_path=None, tenant_id: str = None):
    if report_date is None:
        report_date = date.today()
    if file_path is None:
        file_path = f"daily_report_combined_{tenant_id}.xlsx" if tenant_id else "daily_report_combined.xlsx"

    report_date_str = report_date.strftime("%d-%m-%Y")
    header1 = ["DATE", report_date_str, "", "", "DAILY REPORT", "", "", "", "", "", "", "","", ""]
    # Adjust header1 to match the number of columns in header2
    header2 = ["BATCH", "SHED", "AGE", "OPEN", "MORT", "CULLS", "CLOSING", "TABLE", "JUMBO", "CR", "TOTAL", "HD%", "STD"]

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

    # Find if today's report already exists and remove it
    found_row = None
    for i in range(ws.max_row, 0, -1):  # Start from last row to first
        row = [cell.value for cell in ws[i]]
        if row and row[0] == "DATE" and row[1] == report_date_str:
            found_row = i
            break


    if found_row:
        # Remove existing block (today's DATE to TOTAL row)
        end_row = found_row
        while end_row <= ws.max_row:
            if ws.cell(row=end_row, column=1).value == "TOTAL":
                break
            end_row += 1
        ws.delete_rows(found_row, end_row - found_row + 1)

    # Append data to bottom
    # Check if the last row is empty and add a blank row for spacing
    last_row_values = [cell.value for cell in ws[ws.max_row]]
    if any(last_row_values):  # if the last row is not completely empty
        ws.append([])         # add a blank row for spacing
        start_row = ws.max_row + 2
    else:
        start_row = ws.max_row + 1

    # Now append header and data
    ws.append(header1)
    ws.append(header2)
    
    # Write batch rows
    total_open = total_mort = total_culls = total_closing = total_table = total_jumbo = total_cr = total_total = total_hd_percent_dummy = total_std_dummy = 0
    for batch in batches:
        row = [
            batch.batch_no,
            batch.shed_no,
            batch.age,
            batch.opening_count,
            batch.mortality,
            batch.culls,
            batch.closing_count,
            getattr(batch, "table_eggs", 0),
            getattr(batch, "jumbo", 0),
            getattr(batch, "cr", 0),
            getattr(batch, "total", 0),
            round(batch.closing_count / max(1, batch.opening_count), 4), # Dummy HD% calculation
            round(float(batch.age) * 2.5, 1) if batch.age else round(float(batch.age) * 2.5, 1) # Dummy STD calculation
        ]
        ws.append(row)
        total_open += batch.opening_count
        total_mort += batch.mortality
        total_culls += batch.culls
        total_closing += batch.closing_count
        total_table += getattr(batch, "table_eggs", 0)
        total_jumbo += getattr(batch, "jumbo", 0)
        total_cr += getattr(batch, "cr", 0)
        total_total += getattr(batch, "total", 0)
        # Accumulate dummy values for totals
        total_hd_percent_dummy += round(batch.closing_count / max(1, batch.opening_count), 4)
        total_std_dummy += round(float(batch.age) * 2.5, 1)
        

    # Write totals row
    ws.append([
        "TOTAL", "", 0, total_open, total_mort, total_culls, total_closing,
        total_table, total_jumbo, total_cr, total_total,
         round(total_hd_percent_dummy / len(batches) if batches else 0, 4), # Dummy average HD%
         round(total_std_dummy / len(batches) if batches else 0, 1) # Dummy average STD
    ])

    # Apply formatting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_red_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid") # Closer to the image's orange-red
    bold_font_black = Font(bold=True, color="000000") # Black font for yellow fill
    bold_font_white = Font(bold=True, color="FFFFFF") # White font for red/orange-red fill
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    bold_font = Font(bold=True)

     # Formatting for header1 (DATE and DAILY REPORT)
    date_cell = ws.cell(row=start_row, column=1)
    date_cell.fill = yellow_fill
    date_cell.font = bold_font_black

    report_date_cell = ws.cell(row=start_row, column=2)
    report_date_cell.fill = yellow_fill
    report_date_cell.font = bold_font_black

    daily_report_cell = ws.cell(row=start_row, column=5) # Column E (5th column) for "DAILY REPORT"
    daily_report_cell.value = "DAILY REPORT"
    daily_report_cell.fill = red_fill
    daily_report_cell.font = bold_font_white
    # Merge cells for "DAILY REPORT"
    ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=len(header2))
    daily_report_cell.alignment = Alignment(horizontal='center', vertical='center')


    # Formatting for header2 (BATCH, SHED, etc.)
    for col_idx, cell in enumerate(ws[start_row + 1]):
        cell.fill = orange_red_fill
        cell.font = bold_font_white
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 10 # Adjust width as needed

    # Formatting for Total row
    for cell in ws[ws.max_row]:
        cell.fill = yellow_fill
        cell.font = bold_font_black

    wb.save(file_path)
    return file_path